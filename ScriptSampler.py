# license: GPL v3
import json
import glob
import os
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from fuzzywuzzy import process


def SanitizeName(s):
	s = s.lower()
	s = s.replace(' ', '_')
	s = s.replace("'","")
	if s == "mephit":
		s = "mezepheles"
	s = s.strip()
	return s
	
def WeightedSampleFromDict(d):
	k = np.array(list(d.keys()))
	p = np.array(list(d.values())) / np.sum(list(d.values()))
	return np.random.choice(k, p=p)
	

# a version of Standard "Amy" Order
standardAmyOrder = [
	"you start",
	"each night,",
	"each night*",
	"each day",
	"once per game, at night,",
	"once per game, at night*",
	"once per game, during the day",
	"once per game",
	"on your 1st night",
	"on your 1st day",
	"when",
	"if you",
	"if",
	"you",
]

def SAO(_s):
	s = _s.lower()
	for i,prefix in enumerate(standardAmyOrder):
		if s.startswith(prefix):
			return i
	return len(standardAmyOrder)	

class Data:
	def __init__(self, path):
		self.path = path
		
		self.teams = {} # teamName -> teamList
		self.roles = [] # idx -> role
		self.rolesInv = {} # role -> idx
		self.roleSAOs = {} # role -> saoClass
		self.roleTeams = {} # role -> team
		self.LoadRoles()
		
		self.jinxes = [] # idx -> (role1, role2)
		self.LoadJinxes()
		
		self.roleAdjacency = None # [roleIdx1][roleIdx2] -> weight
		self.saoDist = {} # saoClass -> weight	 
		self.LoadScripts()
		
		self.hardRestrictions = {} # role1 -> requires role2
		self.LoadHardRestrictions()

		self.WriteStats()
		
	def LoadRoles(self):
		teamNames = ["townsfolk", "outsider", "minion", "demon"]
		for team in teamNames:
			self.teams[team] = []

		with open(os.path.join(self.path, "official", "roles.json")) as j:
			jsonRoles = json.load(j)
			jsonRoles = sorted(jsonRoles, key=lambda x: SAO(x["ability"]))
			jsonRoles = sorted(jsonRoles, key=lambda x: x["team"], reverse=True)
			for role in jsonRoles:
				if role["team"] not in self.teams:
					continue
				if role["id"] == "mephit":
					continue # included with both names in roles.json
				roleId = SanitizeName(role["name"])
				self.teams[role["team"]].append(roleId)
				self.roles.append(roleId)
				self.rolesInv[roleId] = len(self.roles)-1
				self.roleSAOs[roleId] = SAO(role["ability"])
				self.roleTeams[roleId] = role["team"]				
	
	def LoadJinxes(self):
		with open(os.path.join(self.path, "official", "hatred.json")) as j:
			jsonHatred = json.load(j)
			for jinx in jsonHatred:
				char1 = SanitizeName(jinx["id"])
				for hatred in jinx["hatred"]:
					char2 = SanitizeName(hatred["id"])
					self.jinxes.append((char1, char2))	
	
	def LoadScripts(self):
		for i in sorted(self.roleSAOs.values()):
			self.saoDist[i] = 0

		self.roleAdjacency = np.zeros((len(self.roles), len(self.roles)))
		scripts = glob.glob(os.path.join(self.path,"scripts/") + "*.json")
		for script in scripts:
			with open(script) as j:
				try:
					jsonScript = json.load(j)
				except:
					continue
				scriptRoles = []
				for role in jsonScript:
					roleId = SanitizeName(role["id"])
					if roleId not in self.roles:
						continue
					scriptRoles.append(roleId)
					assert(roleId in self.roles)
					if roleId in self.teams["townsfolk"]:
						self.saoDist[self.roleSAOs[roleId]] += 1
				for role1 in scriptRoles:
					for role2 in scriptRoles:
						if role1 == role2:
							continue
						self.roleAdjacency[self.rolesInv[role1],self.rolesInv[role2]] += 1
		
	def LoadHardRestrictions(self):
		with open(os.path.join(self.path,"hardRestrictions.json")) as f:
			self.hardRestrictions = json.load(f)
	
	def WriteStats(self):
		# heatmap of adjacency matrix
		plt.clf()
		plt.figure(figsize = (10,10))
		plt.imshow(self.roleAdjacency, cmap='hot', interpolation="antialiased")
		dividers = [len(self.teams["townsfolk"]), 
					len(self.teams["townsfolk"])+len(self.teams["outsider"]), 
					len(self.teams["townsfolk"])+len(self.teams["outsider"])+len(self.teams["minion"])]
		for v in dividers:
			plt.axvline(x=v,color='green')
			plt.axhline(y=v,color='green')
		plt.gca().invert_yaxis()
		plt.ylabel("role")
		plt.xlabel("role")
		plt.savefig(os.path.join(self.path,"stats","heatmap.png"), bbox_inches="tight")		   

		# write heatmap as xlsx, with role names included
		file_name = 'heatmap.xlsx'
		wb = openpyxl.Workbook()
		ws = wb.worksheets[0]
		ws.title = 'heatmap'
		for row, rowVals in enumerate(self.roleAdjacency):
			for col, item in enumerate(rowVals):
				ws.cell(column=col+2, row=row+2, value=item)
				r = 255 * item / np.max(self.roleAdjacency)
				hexColour = '%02x%02x%02x' % (255, 255-int(r), 255-int(r))
				ws.cell(column=col+2, row=row+2).fill = openpyxl.styles.PatternFill(start_color=hexColour, end_color=hexColour, fill_type = "solid")
		for i,role in enumerate(self.roles):
			ws.cell(column=1, row=i+2, value=role)
			ws.cell(column=i+2, row=1, value=role)
		wb.save(os.path.join(self.path, "stats", file_name))

		# SAO distribution
		plt.clf()
		keys = []
		for k in self.saoDist.keys():
			if k < len(standardAmyOrder):
				keys.append("\n".join(standardAmyOrder[k].split()))
			else:
				keys.append("other")
		plt.bar(keys, self.saoDist.values())
		plt.ylabel("frequency")
		plt.savefig(os.path.join(self.path,"stats","sao.png"), bbox_inches="tight")		
		plt.clf()
		

class Script:
	def __init__(self, inputData, teamSizes, seed=0, steps=1000, alpha=0, beta=1, gamma=0.1, requiredRoles=[], omittedRoles=[]):
		self.data = inputData
		self.teamSizes = teamSizes
		self.seed = seed
		self.nSteps = steps
		self.alpha = alpha # pref attach init weight
		self.beta = beta # pref attach power
		self.gamma = gamma # distance along role weights vector to use in sampling (e.g. 0.5=median)
		np.random.seed(seed)
		
		self.requiredRoles = set()
		self.omittedRoles = set()
		for roleArg in requiredRoles:
			self.requiredRoles.add(process.extractOne(roleArg, self.data.roles)[0])
		for roleArg in omittedRoles:
			self.omittedRoles.add(process.extractOne(roleArg, self.data.roles)[0])
		self.omittedRoles = self.omittedRoles - self.requiredRoles
		
		self.script = {} # team -> [roles]
		for team,n in self.teamSizes.items():
			rolesInit = list(set(self.data.teams[team]) - self.omittedRoles)
			self.script[team] = np.random.choice(rolesInit, n, replace=False)

		# insert required roles
		for role in self.requiredRoles:
			team = self.data.roleTeams[role]
			if role in self.script[team]:
				continue
			for i in range(len(self.script[team])):
				if self.script[team][i] in self.requiredRoles:
					continue
				self.script[team][i] = role
				break
		
		self.BuildScript()
			
	def ListRoles(self):
		return [role for team in self.script for role in self.script[team]]
	
	def BuildScript(self):
		n = 0
		while ((self.IsTheScriptActuallyBroken() or n < self.nSteps) and n < self.nSteps*2):
			self.Step()
			n += 1
			
	def Step(self):
		# Gibbs sampler step
		
		# choose which slot to resample & empty it
		valid = False
		for tries in range(20):
			team = WeightedSampleFromDict(self.teamSizes)
			pos = np.random.randint(0,len(self.script[team]))
			valid = self.script[team][pos] not in self.requiredRoles
		if not valid:
			return		
		self.script[team][pos] = ""

		# set role weights according to (sum of) adjacency to current roles
		sao = WeightedSampleFromDict(self.data.saoDist) # filter by SAO dist if townsfolk
		scriptRoles = self.ListRoles()
		roleWeightsFull = {}
		for role1 in scriptRoles:
			if not role1:
				continue
			for role2 in self.data.roles:
				if self.data.roleTeams[role2] != team:
					continue
				if role2 in scriptRoles:
					continue
				if role2 in self.omittedRoles:
					continue
				if role1 == role2:
					continue
				if role2 not in roleWeightsFull:
					roleWeightsFull[role2] = []
				roleWeightsFull[role2].append(self.data.roleAdjacency[self.data.rolesInv[role1],self.data.rolesInv[role2]])
		
		# record the gamma'th percentile of roleWeightsFull for each role
		# use that as sampling weight in pref-attach style sampler
		roleWeights = {}
		for role in roleWeightsFull:
			roleWeightsFull[role] = sorted(roleWeightsFull[role])
			w = np.percentile(roleWeightsFull[role], self.gamma*100)
			roleWeights[role] = self.alpha + w ** self.beta

		if np.sum(list(roleWeights.values())) > 0:
			newRole = WeightedSampleFromDict(roleWeights)
		else:
			newRole = np.random.choice(list(set(self.data.teams[team]) - set(scriptRoles)))
		self.script[team][pos] = newRole				  

		# debugging
		#print(scriptRoles)
		#print(json.dumps(roleWeights, indent=2, sort_keys=True))
		
	def IsTheScriptActuallyBroken(self):
		scriptRoles = self.ListRoles()
		for k,v in self.data.hardRestrictions.items():
			if k in scriptRoles and v not in scriptRoles:
				return True
		if self.CountJinxes() > 5:
			return True
		return False
	
	def CountJinxes(self):
		nJinxes = 0
		scriptRoles = self.ListRoles()
		for (role1,role2) in self.data.jinxes:
			if role1 in scriptRoles and role2 in scriptRoles:
				nJinxes += 1
		return nJinxes
	
	def ScriptRoleAffinity(self):
		# normalized (to 1) average heatmap value between script roles
		roleWeights = {}
		scriptRoles = self.ListRoles()
		for role1 in scriptRoles:
			if not role1:
				continue
			for role2 in self.data.roles:
				if role2 not in roleWeights:
					roleWeights[role2] = self.alpha * np.median(self.data.roleAdjacency) 
				roleWeights[role2] += self.data.roleAdjacency[self.data.rolesInv[role1],self.data.rolesInv[role2]]
		return np.sum(list(roleWeights.values())) / len(scriptRoles) / np.median(self.data.roleAdjacency) / len(self.data.roleAdjacency)
	
	def ScriptSAODist(self):
		saoDist = {}
		scriptRoles = self.ListRoles()
		for role in scriptRoles:
			if not role:
				continue
			if not self.data.roleTeams[role] == "townsfolk":
				continue
			saoClass = self.data.roleSAOs[role]
			if saoClass not in saoDist:
				saoDist[saoClass] = 0
			saoDist[saoClass] += 1
		return saoDist 
		
	def ScriptSAOVariation(self):
		# total variation of current SAO dist from data distribution
		saoDist = self.ScriptSAODist()
		n = max(max(saoDist.keys()), max(self.data.saoDist.keys()))+1 
		v1 = np.array([saoDist.get(i,0) for i in range(n)]) / np.sum(list(saoDist.values()))
		v2 = np.array([self.data.saoDist.get(i,0) for i in range(n)]) / np.sum(list(self.data.saoDist.values()))
		return np.sum(np.abs(v1-v2))
	
	def SAOsort(self, roleList):
		roleList = sorted(roleList)
		roleList = sorted(roleList, key=lambda x: self.data.roleSAOs[x])
		roleList = sorted(roleList, key=lambda x: self.data.roleTeams[x], reverse=True)
		return roleList
	
	def __repr__(self):
		s = ''
		if len(self.requiredRoles) > 0:
			s += '**Required Roles:**  ' + '  '.join(self.requiredRoles) + '\n'
		if len(self.omittedRoles) > 0:
			s += '**Omitted Roles:**  ' + '  '.join(self.omittedRoles) + '\n'
		for team,teamRoles in self.script.items():
			teamRoles = self.SAOsort(teamRoles)
			s += '**' + team.title() + ':**  '
			s += '  '.join(teamRoles)
			s += '\n'
		s += 'jinxes: %d' % self.CountJinxes()
		s += '\nid: %s' % self.ID()
		if self.IsTheScriptActuallyBroken():
			s += '\nFailed validity check!'
		return s
	
	def ID(self):
		return "%d_%d_%.1f_%.1f_%.1f_%.2f_%.2f" % (self.seed, self.nSteps, self.alpha, self.beta, self.gamma, self.ScriptRoleAffinity(), self.ScriptSAOVariation())
	
	def ToolScript(self):
		j = []
		roles = self.SAOsort(self.ListRoles())
		for role in roles:
			j.append({"id":role})
		return j
	
	def Save(self):
		# save in json format compatible with clocktower.online
		with open("script_"+self.ID()+'.json', 'w') as f:
			json.dump(self.ToolScript(), f)
			
	
			   